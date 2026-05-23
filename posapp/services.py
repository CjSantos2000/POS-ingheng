from __future__ import annotations

import csv
import io
import secrets
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal

from openpyxl import load_workbook
from reportlab.graphics.barcode import code128
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

from django.conf import settings
from django.core.cache import cache
from django.db import transaction
from django.db.models import DecimalField, ExpressionWrapper, F, Sum
from django.http import HttpResponse
from django.utils import timezone

from .models import Product, SaleItem, SaleTransaction, StockMovement, reduce_product_stock


CART_CACHE_TIMEOUT = 60 * 60 * 8


@dataclass
class CartTotals:
    subtotal: Decimal
    tax: Decimal
    total: Decimal


@dataclass
class ImportSummary:
    created: int = 0
    updated: int = 0
    generated_barcodes: int = 0
    stock_adjustments: int = 0
    failed_rows: int = 0
    errors: list[str] | None = None

    def __post_init__(self):
        if self.errors is None:
            self.errors = []


def get_cart_cache_key(user_id: int) -> str:
    return f"pos-cart-{user_id}"


def get_cart(user_id: int) -> dict:
    return cache.get(get_cart_cache_key(user_id), {})


def save_cart(user_id: int, cart: dict) -> None:
    cache.set(get_cart_cache_key(user_id), cart, timeout=CART_CACHE_TIMEOUT)


def clear_cart(user_id: int) -> None:
    cache.delete(get_cart_cache_key(user_id))


def calculate_cart_totals(cart: dict, discount_amount: Decimal = Decimal("0.00")) -> CartTotals:
    subtotal = sum(Decimal(str(item["subtotal"])) for item in cart.values())
    tax = (subtotal - discount_amount) * settings.POS_TAX_RATE
    total = max(subtotal - discount_amount + tax, Decimal("0.00"))
    return CartTotals(subtotal=subtotal, tax=tax.quantize(Decimal("0.01")), total=total.quantize(Decimal("0.01")))


def generate_product_sku(name: str) -> str:
    base = "".join(character for character in name.upper() if character.isalnum())[:8] or "ITEM"
    while True:
        candidate = f"{base}{secrets.randbelow(10000):04d}"
        if not Product.objects.filter(sku=candidate).exists():
            return candidate


def generate_unique_barcode() -> str:
    while True:
        candidate = f"27{secrets.randbelow(10**11):011d}"
        if not Product.objects.filter(barcode=candidate).exists():
            return candidate


def resolve_product_for_import(*, barcode: str, sku: str) -> Product | None:
    product = None
    if barcode:
        product = Product.objects.filter(barcode=barcode).first()
    if product is None and sku:
        product = Product.objects.filter(sku=sku).first()
    return product


def parse_decimal(value, default: str = "0") -> Decimal:
    if value in (None, ""):
        value = default
    return Decimal(str(value))


def parse_int(value, default: int = 0) -> int:
    if value in (None, ""):
        value = default
    return int(value)


def normalize_category(value: str) -> str:
    cleaned = str(value or "").strip()
    if not cleaned:
        return Product.Category.OTHER

    for option in Product.Category:
        if cleaned == option.value or cleaned.lower() == option.label.lower():
            return option.value
    return Product.Category.OTHER


def scan_product_into_cart(*, user_id: int, barcode: str) -> tuple[dict, CartTotals]:
    # Only pull the columns needed for scan-time response to keep barcode lookups fast.
    product = Product.objects.only("id", "name", "barcode", "selling_price", "stock_quantity").get(barcode=barcode, is_active=True)
    cart = get_cart(user_id)
    item = cart.get(str(product.id))

    # Check stock before adding to cart
    current_qty = item["quantity"] if item else 0
    new_qty = current_qty + 1
    
    if new_qty > product.stock_quantity:
        raise ValueError(
            f"Cannot add {product.name} - only {product.stock_quantity} in stock "
            f"(cart already has {current_qty})"
        )

    if item:
        item["quantity"] += 1
        item["subtotal"] = str((Decimal(item["unit_price"]) * item["quantity"]).quantize(Decimal("0.01")))
    else:
        item = {
            "product_id": product.id,
            "barcode": product.barcode,
            "name": product.name,
            "unit_price": str(product.selling_price),
            "quantity": 1,
            "subtotal": str(product.selling_price.quantize(Decimal("0.01"))),
            "stock_quantity": product.stock_quantity,
        }
        cart[str(product.id)] = item

    save_cart(user_id, cart)
    totals = calculate_cart_totals(cart)
    return item, totals


def update_cart_item_quantity(*, user_id: int, product_id: int, quantity: int) -> CartTotals:
    cart = get_cart(user_id)
    key = str(product_id)
    if key not in cart:
        return calculate_cart_totals(cart)
    if quantity <= 0:
        cart.pop(key)
    else:
        unit_price = Decimal(cart[key]["unit_price"])
        cart[key]["quantity"] = quantity
        cart[key]["subtotal"] = str((unit_price * quantity).quantize(Decimal("0.01")))
    save_cart(user_id, cart)
    return calculate_cart_totals(cart)


def cart_snapshot(user_id: int) -> tuple[list[dict], CartTotals]:
    cart = get_cart(user_id)
    items = list(cart.values())
    return items, calculate_cart_totals(cart)


def generate_receipt_number() -> str:
    return timezone.localtime().strftime("R%Y%m%d%H%M%S%f")[-18:]


@transaction.atomic
def finalize_sale(*, user, customer, payment_method: str, discount_amount: Decimal) -> SaleTransaction:
    cart = get_cart(user.id)
    if not cart:
        raise ValueError("Cart is empty.")

    product_ids = [int(product_id) for product_id in cart.keys()]
    # Lock scanned rows only while checkout is running so concurrent cashiers do not oversell.
    products = Product.objects.select_for_update().filter(id__in=product_ids, is_active=True)
    product_map = {product.id: product for product in products}

    for key, item in cart.items():
        product = product_map.get(int(key))
        if product is None:
            raise ValueError(f"Product {key} is unavailable.")
        if product.stock_quantity < item["quantity"]:
            raise ValueError(f"Insufficient stock for {product.name}.")

    totals = calculate_cart_totals(cart, discount_amount=discount_amount)
    transaction_obj = SaleTransaction.objects.create(
        receipt_number=generate_receipt_number(),
        cashier=user,
        customer=customer,
        subtotal=totals.subtotal,
        tax_amount=totals.tax,
        discount_amount=discount_amount,
        total_amount=totals.total,
        payment_method=payment_method,
    )

    sale_items = []
    stock_movements = []
    for key, item in cart.items():
        product_id = int(key)
        quantity = item["quantity"]
        unit_price = Decimal(item["unit_price"])
        line_total = Decimal(item["subtotal"])
        sale_items.append(
            SaleItem(
                transaction=transaction_obj,
                product_id=product_id,
                quantity=quantity,
                unit_price=unit_price,
                line_total=line_total,
            )
        )
        reduce_product_stock(product_id, quantity)
        stock_movements.append(
            StockMovement(
                product_id=product_id,
                movement_type=StockMovement.MovementType.SALE,
                quantity_delta=-quantity,
                reference=transaction_obj.receipt_number,
            )
        )

    SaleItem.objects.bulk_create(sale_items)
    StockMovement.objects.bulk_create(stock_movements)
    clear_cart(user.id)
    return transaction_obj


def parse_product_upload(uploaded_file) -> list[dict]:
    suffix = uploaded_file.name.lower()
    if suffix.endswith(".csv"):
        content = uploaded_file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        return [row for row in reader]

    workbook = load_workbook(uploaded_file, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    parsed = []
    for row in rows[1:]:
        parsed.append({headers[index]: value for index, value in enumerate(row)})
    return parsed


@transaction.atomic
def import_products(rows: list[dict], *, generate_missing_barcodes: bool = True) -> ImportSummary:
    summary = ImportSummary()
    for index, row in enumerate(rows, start=2):
        try:
            barcode = str(row.get("barcode", "") or "").strip()
            sku = str(row.get("sku", "") or "").strip()
            name = str(row.get("name", "Unnamed Item") or "Unnamed Item").strip()

            product = resolve_product_for_import(barcode=barcode, sku=sku)
            previous_stock = product.stock_quantity if product else 0

            generated_now = False
            if not barcode:
                if product and product.barcode:
                    barcode = product.barcode
                elif generate_missing_barcodes:
                    barcode = generate_unique_barcode()
                    generated_now = True
                else:
                    raise ValueError("barcode is missing")

            if not sku:
                sku = product.sku if product and product.sku else generate_product_sku(name or barcode)

            if product is None:
                product = Product()
                summary.created += 1
            else:
                summary.updated += 1

            if Product.objects.exclude(pk=product.pk).filter(barcode=barcode).exists():
                raise ValueError(f"barcode {barcode} already belongs to another product")
            if Product.objects.exclude(pk=product.pk).filter(sku=sku).exists():
                raise ValueError(f"SKU {sku} already belongs to another product")

            barcode_was_blank = not str(row.get("barcode", "") or "").strip()
            product.sku = sku
            product.barcode = barcode
            if generated_now:
                product.barcode_source = Product.BarcodeSource.GENERATED
            elif barcode_was_blank and product.pk:
                product.barcode_source = product.barcode_source
            else:
                product.barcode_source = Product.BarcodeSource.MANUAL
            product.name = name
            product.category = normalize_category(row.get("category", product.category or Product.Category.OTHER))
            product.description = str(row.get("description", product.description or "") or "").strip()
            product.cost_price = parse_decimal(row.get("cost_price", product.cost_price or "0"))
            product.selling_price = parse_decimal(row.get("selling_price", product.selling_price or "0"))
            product.stock_quantity = parse_int(row.get("stock_quantity", product.stock_quantity if product.pk else 0), product.stock_quantity if product.pk else 0)
            product.reorder_level = parse_int(row.get("reorder_level", product.reorder_level or settings.POS_LOW_STOCK_DEFAULT), settings.POS_LOW_STOCK_DEFAULT)
            product.is_active = str(row.get("is_active", product.is_active if product.pk else "true")).lower() not in {"false", "0", "no"}
            product.full_clean()
            product.save()

            stock_delta = product.stock_quantity - previous_stock
            if stock_delta:
                StockMovement.objects.create(
                    product=product,
                    movement_type=StockMovement.MovementType.IMPORT,
                    quantity_delta=stock_delta,
                    reference="catalog-import",
                )
                summary.stock_adjustments += 1

            if generated_now:
                summary.generated_barcodes += 1
        except Exception as exc:
            summary.failed_rows += 1
            summary.errors.append(f"Row {index}: {exc}")
    return summary


@transaction.atomic
def import_stock_rows(rows: list[dict], *, reference: str) -> ImportSummary:
    summary = ImportSummary()
    for index, row in enumerate(rows, start=2):
        try:
            barcode = str(row.get("barcode", "") or "").strip()
            sku = str(row.get("sku", "") or "").strip()
            quantity = parse_int(row.get("stock_in_quantity", 0), 0)
            movement_reference = str(row.get("reference", "") or reference).strip() or reference

            if quantity <= 0:
                raise ValueError("stock_in_quantity must be greater than zero")

            product = resolve_product_for_import(barcode=barcode, sku=sku)
            if product is None:
                raise ValueError("product was not found by barcode or SKU")

            product.stock_quantity = product.stock_quantity + quantity
            product.save(update_fields=["stock_quantity", "updated_at"])
            StockMovement.objects.create(
                product=product,
                movement_type=StockMovement.MovementType.IMPORT,
                quantity_delta=quantity,
                reference=movement_reference,
            )
            summary.updated += 1
            summary.stock_adjustments += quantity
        except Exception as exc:
            summary.failed_rows += 1
            summary.errors.append(f"Row {index}: {exc}")
    return summary


def export_products_csv() -> HttpResponse:
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="products_export.csv"'
    writer = csv.writer(response)
    writer.writerow(["sku", "barcode", "barcode_source", "name", "category", "description", "cost_price", "selling_price", "stock_quantity", "reorder_level", "is_active"])
    for product in Product.objects.order_by("name").values_list(
        "sku", "barcode", "barcode_source", "name", "category", "description", "cost_price", "selling_price", "stock_quantity", "reorder_level", "is_active"
    ):
        writer.writerow(product)
    return response


def export_stock_template_csv() -> HttpResponse:
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="stock_in_template.csv"'
    writer = csv.writer(response)
    writer.writerow(["barcode", "sku", "stock_in_quantity", "reference"])
    writer.writerow(["2700000000001", "NOTE0001", "12", "delivery-2026-05-07"])
    return response


def export_stock_levels_csv() -> HttpResponse:
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="stock_levels_export.csv"'
    writer = csv.writer(response)
    writer.writerow(["sku", "barcode", "name", "stock_quantity", "reorder_level", "is_active"])
    for product in Product.objects.order_by("name").values_list("sku", "barcode", "name", "stock_quantity", "reorder_level", "is_active"):
        writer.writerow(product)
    return response


def build_receipt_pdf(transaction_obj: SaleTransaction) -> HttpResponse:
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{transaction_obj.receipt_number}.pdf"'
    currency = settings.POS_CURRENCY_SYMBOL

    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    y = height - 50

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(50, y, "POS Enghing Receipt")
    y -= 25
    pdf.setFont("Helvetica", 11)
    pdf.drawString(50, y, f"Receipt: {transaction_obj.receipt_number}")
    y -= 18
    pdf.drawString(50, y, f"Date: {timezone.localtime(transaction_obj.created_at).strftime('%Y-%m-%d %H:%M:%S')}")
    y -= 18
    pdf.drawString(50, y, f"Cashier: {transaction_obj.cashier.username}")
    y -= 28

    for item in transaction_obj.items.select_related("product"):
        pdf.drawString(50, y, f"{item.product.name} x {item.quantity}")
        pdf.drawRightString(width - 50, y, f"{currency}{item.line_total}")
        y -= 18

    y -= 10
    pdf.drawString(50, y, f"Subtotal: {currency}{transaction_obj.subtotal}")
    y -= 18
    pdf.drawString(50, y, f"Tax: {currency}{transaction_obj.tax_amount}")
    y -= 18
    pdf.drawString(50, y, f"Discount: {currency}{transaction_obj.discount_amount}")
    y -= 18
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(50, y, f"Total: {currency}{transaction_obj.total_amount}")

    pdf.showPage()
    pdf.save()
    return response


def build_thermal_receipt_pdf(transaction_obj: SaleTransaction) -> HttpResponse:
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{transaction_obj.receipt_number}-thermal.pdf"'
    currency = settings.POS_CURRENCY_SYMBOL

    items = list(transaction_obj.items.select_related("product"))
    page_width = 80 * mm
    row_height = 6.5 * mm
    base_height = 75 * mm
    computed_height = base_height + (len(items) * row_height) + (12 * row_height)
    page_height = max(computed_height, 140 * mm)

    pdf = canvas.Canvas(response, pagesize=(page_width, page_height))
    y = page_height - 8 * mm
    left = 4 * mm
    right = page_width - 4 * mm

    def divider():
        nonlocal y
        pdf.setLineWidth(0.4)
        pdf.line(left, y, right, y)
        y -= 3.5 * mm

    def row(label: str, value: str, *, bold: bool = False):
        nonlocal y
        pdf.setFont("Courier-Bold" if bold else "Courier", 10)
        pdf.drawString(left, y, label)
        pdf.drawRightString(right, y, value)
        y -= row_height

    pdf.setFont("Courier-Bold", 12)
    pdf.drawCentredString(page_width / 2, y, "POS Enghing")
    y -= row_height
    pdf.setFont("Courier", 10)
    pdf.drawCentredString(page_width / 2, y, "Retail Invoice")
    y -= row_height

    divider()
    row("Date", timezone.localtime(transaction_obj.created_at).strftime("%d/%m/%Y %I:%M %p"))
    row("Cashier", transaction_obj.cashier.username)
    if transaction_obj.customer:
        row("Customer", transaction_obj.customer.name[:22])
    row("Bill No", transaction_obj.receipt_number)
    row("Payment", transaction_obj.get_payment_method_display())
    divider()

    pdf.setFont("Courier-Bold", 10)
    pdf.drawString(left, y, "Item")
    pdf.drawRightString(right - 18 * mm, y, "Qty")
    pdf.drawRightString(right, y, "Amt")
    y -= row_height
    divider()

    total_qty = 0
    for item in items:
        total_qty += item.quantity
        name = item.product.name[:22]
        pdf.setFont("Courier", 10)
        pdf.drawString(left, y, name)
        pdf.drawRightString(right - 18 * mm, y, str(item.quantity))
        pdf.drawRightString(right, y, f"{item.line_total:.2f}")
        y -= row_height

    divider()
    row("Sub Total", f"{currency}{transaction_obj.subtotal:.2f}", bold=True)
    row("Items", str(total_qty))
    row("Tax", f"{currency}{transaction_obj.tax_amount:.2f}")
    row("Discount", f"{currency}{transaction_obj.discount_amount:.2f}")
    divider()
    row("TOTAL", f"{currency}{transaction_obj.total_amount:.2f}", bold=True)
    divider()
    row("Cash", f"{currency}{transaction_obj.total_amount:.2f}")
    row("Cash tendered", f"{currency}{transaction_obj.total_amount:.2f}")

    y -= 2 * mm
    pdf.setFont("Courier", 9)
    pdf.drawCentredString(page_width / 2, y, "THANK YOU. PLEASE COME AGAIN")

    pdf.showPage()
    pdf.save()
    return response


def build_product_barcodes_pdf(products, *, labels_per_product: int = 1) -> HttpResponse:
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="product-barcodes.pdf"'

    pdf = canvas.Canvas(response, pagesize=A4)
    page_width, page_height = A4
    label_width = 60 * mm
    label_height = 30 * mm
    left_margin = 12 * mm
    top_margin = 15 * mm
    horizontal_gap = 6 * mm
    vertical_gap = 6 * mm
    columns = 3

    x = left_margin
    y = page_height - top_margin - label_height
    column = 0

    for product in products:
        for _ in range(max(1, labels_per_product)):
            pdf.roundRect(x, y, label_width, label_height, 3 * mm, stroke=1, fill=0)
            pdf.setFont("Helvetica-Bold", 9)
            pdf.drawString(x + 3 * mm, y + label_height - 6 * mm, product.name[:28])
            pdf.setFont("Helvetica", 7)
            pdf.drawString(x + 3 * mm, y + label_height - 10 * mm, f"SKU: {product.sku}")

            barcode = code128.Code128(product.barcode, barHeight=10 * mm, barWidth=0.4)
            barcode.drawOn(pdf, x + 3 * mm, y + 8 * mm)
            pdf.setFont("Helvetica", 8)
            pdf.drawCentredString(x + (label_width / 2), y + 4 * mm, product.barcode)

            column += 1
            if column == columns:
                column = 0
                x = left_margin
                y -= label_height + vertical_gap
            else:
                x += label_width + horizontal_gap

            if y < 15 * mm:
                pdf.showPage()
                x = left_margin
                y = page_height - top_margin - label_height
                column = 0

    pdf.save()
    return response


def sales_summary_queryset(start_date=None, end_date=None):
    queryset = SaleTransaction.objects.all()
    if start_date:
        queryset = queryset.filter(created_at__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(created_at__date__lte=end_date)
    return queryset


def report_metrics(start_date=None, end_date=None) -> dict:
    queryset = sales_summary_queryset(start_date, end_date)
    gross_sales = queryset.aggregate(total=Sum("total_amount"))["total"] or Decimal("0.00")
    transaction_count = queryset.count()
    items_sold = SaleItem.objects.filter(transaction__in=queryset).aggregate(total=Sum("quantity"))["total"] or 0
    top_products = (
        SaleItem.objects.filter(transaction__in=queryset)
        .values("product__name")
        .annotate(quantity_sold=Sum("quantity"), revenue=Sum("line_total"))
        .order_by("-quantity_sold")[:10]
    )
    return {
        "gross_sales": gross_sales,
        "transaction_count": transaction_count,
        "items_sold": items_sold,
        "top_products": top_products,
    }
